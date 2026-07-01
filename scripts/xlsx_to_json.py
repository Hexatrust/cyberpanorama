#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Sens INVERSE : reinjecte les modifications de l'Excel dans les JSON sources.

Lit data/solutions_master.xlsx (feuille 'Solutions') et reecrit, par id de solution :
  - data/solutions.json    : le fichier unique de donnee (nom, description, NIST, site, contact, taille,
                             indexation = mots-cles de recherche, logo...)
  - data/size_review.json  : code taille (la taille re-verifiee, que le build applique)

NB Logo : la colonne 'Logo' (nom de fichier, ex. cryptr.png) va dans solutions.json ; le fichier image
doit exister dans assets/logos/ (on ne peut pas embarquer une image dans l'Excel).

Apres ce script, relancer build_app_data.py et build_master_xlsx.py pour propager.

Appariement par la colonne 'ID' (ajoutee par build_master_xlsx.py). A defaut, repli par nom d'entreprise.

Tous les champs affiches sur le site font l'aller-retour : aucun n'est verrouille.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

DATA = Path(__file__).resolve().parents[1] / "data"
XLSX = DATA / "solutions_master.xlsx"
SIZE_REVIEW = DATA / "size_review.json"
SOLUTIONS = DATA / "solutions.json"

# Libelle Excel -> code taille interne (inverse de SIZE_FR du build).
SIZE_CODE = {
    "Grand groupe": "very_large",
    "ETI / scale-up": "large",
    "PME": "medium",
    "Startup / petite entreprise": "small",
}

def load(path, default):
    if path.exists():
        with path.open(encoding="utf-8") as fh:
            return json.load(fh) or default
    return default


def clean(v):
    return ("" if v is None else str(v)).strip()


def parse_code(cell):
    """'PR.DS : libelle' -> 'PR.DS' ; 'Protéger' -> 'Protéger' ; '' -> ''."""
    s = clean(cell)
    if not s:
        return ""
    # On garde la partie avant ' : ' (le code), sinon la valeur telle quelle.
    return s.split(" : ", 1)[0].strip().split(":", 1)[0].strip() if " :" in s or ":" in s else s


def parse_name_after(cell):
    """'PR : Protéger' -> 'Protéger' (on garde le libelle apres ' : ', sinon la valeur)."""
    s = clean(cell)
    if " : " in s:
        return s.split(" : ", 1)[1].strip()
    return s


def parse_l3(cell):
    """Multi-lignes 'PR.DS-01 : libelle' -> ['PR.DS-01', ...]."""
    out = []
    for line in clean(cell).splitlines():
        code = parse_code(line)
        if code:
            out.append(code)
    return out


def main():
    if not XLSX.exists():
        sys.exit(f"Excel introuvable : {XLSX}")
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if "Solutions" not in wb.sheetnames:
        sys.exit("Feuille 'Solutions' absente du classeur.")
    ws = wb["Solutions"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Feuille 'Solutions' vide.")
    header = [clean(h) for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    size_review = load(SIZE_REVIEW, {})

    # solutions.json = liste d'acteurs ; on indexe par id pour modifier les entrees en place.
    sols = load(SOLUTIONS, [])
    by_id = {s.get("id"): s for s in sols}

    # Repli d'appariement par nom si pas de colonne ID.
    name_to_id = {}
    if "ID" not in idx:
        for s in sols:
            name_to_id[clean(s.get("solution_name")).lower()] = s.get("id")

    warns, updated = [], 0
    for r in rows[1:]:
        def col(h):
            return r[idx[h]] if h in idx and idx[h] < len(r) else None

        sid = clean(col("ID")) if "ID" in idx else name_to_id.get(clean(col("Entreprise")).lower(), "")
        if not sid:
            warns.append(f"ligne sans id (Entreprise={clean(col('Entreprise'))!r}) ignoree")
            continue
        v = by_id.get(sid)
        if v is None:
            warns.append(f"{sid}: id absent de solutions.json, ligne ignoree")
            continue

        # Nom
        name = clean(col("Entreprise"))
        if name:
            v["solution_name"] = name
            v["company_name"] = name

        # Description (courte affichee) + detaillee (moteur de recherche)
        desc = clean(col("Description"))
        if desc:
            v["description"] = desc
        ddesc = clean(col("Description longue"))
        if ddesc:
            v["detailed_description"] = ddesc

        # Site / contact / NIS2
        web = clean(col("Site web"))
        if web:
            v["website"] = web
        contact = clean(col("Contact"))
        if contact:
            if "@" in contact:
                v["email_contact"] = contact
            else:
                v["contact_url"] = contact
        nis2 = clean(col("Objectif NIS2"))
        if nis2:
            v["nis2_objective"] = nis2

        # NIST
        l1 = parse_name_after(col("NIST 1 (Fonction)"))
        l2 = parse_code(col("NIST 2 (Catégorie)"))
        l3 = parse_l3(col("NIST 3 (Sous-catégories)"))
        if l1 or l2 or l3:
            nist = dict(v.get("nist") or {})
            if l1:
                nist["level1"] = l1
            if l2:
                nist["level2"] = [l2]
            if l3:
                nist["level3"] = l3
            v["nist"] = nist

        # Taille : dans solutions.json ET dans size_review (c'est lui que le build applique).
        size_label = clean(col("Type d'entreprise"))
        if size_label:
            code = SIZE_CODE.get(size_label)
            if not code:
                warns.append(f"{sid}: taille '{size_label}' non reconnue, ignoree")
            else:
                v["size"] = code
                sr = dict(size_review.get(sid) or {})
                sr["size_code"] = code
                size_review[sid] = sr

        # Indexation = mots-cles de recherche (liste separee par des virgules dans l'Excel)
        idx_cell = clean(col("Indexation"))
        if idx_cell:
            v["indexation"] = [k.strip() for k in idx_cell.split(",") if k.strip()]

        # Logo (nom de fichier) : on l'ecrit s'il change. Le fichier image doit exister dans assets/logos/.
        logo = clean(col("Logo"))
        if logo and v.get("logo_file") != logo:
            v["logo_path"] = f"assets/logos/{logo}"
            v["logo_file"] = logo
            v["logo_source"] = "excel"

        updated += 1

    SOLUTIONS.write_text(json.dumps(sols, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    SIZE_REVIEW.write_text(json.dumps(size_review, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"xlsx -> json : {updated} lignes appliquees")
    if warns:
        print(f"  {len(warns)} avertissement(s) :")
        for w in warns[:40]:
            print(f"   - {w}")


if __name__ == "__main__":
    main()
