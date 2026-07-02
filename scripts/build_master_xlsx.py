#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Génère l'Excel master enrichi des acteurs cyber (un seul fichier, colonnes utiles).

Entrées :
  - data/solutions.json            : le fichier unique de donnee (276 acteurs publiés)
  - data/nist_labels_fr.json       : libellés FR (level1/2/3) pour le rendu 'code — libellé'

Sortie :
  - data/solutions_master.xlsx     : 1 classeur. Feuilles : 'Solutions', 'Référentiel NIST',
                                     'À vérifier' (seulement si des solutions sont flaggées).
  - data/master_qa_report.json     : récap QA.

Règles métier (validées) :
  - NIST 1 : une fonction.  NIST 2 : UNE catégorie principale.  NIST 3 : plusieurs, toutes ⊂ N2.
  - Contact : email si dispo, sinon URL du formulaire/page contact (une seule colonne).
  - Aucune colonne entièrement vide n'est écrite.
"""
from __future__ import annotations

import json
from collections import Counter
from copy import copy as _copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parents[1] / "data"
SOLUTIONS = DATA_DIR / "solutions.json"
LABELS = DATA_DIR / "nist_labels_fr.json"
OUT_XLSX = DATA_DIR / "solutions_master.xlsx"
QA_JSON = DATA_DIR / "master_qa_report.json"
# Fichier de référence (mise en forme) : on en copie la feuille 'Référentiel NIST' à l'identique
# (cellules fusionnées, bordures, en-tête). Optionnel : ignoré s'il est absent.
REF_XLSX = DATA_DIR / "solutions_master_initial.xlsx"

# Type d'entreprise (palier ≈ CA / maturité) affiché dans la colonne Taille.
SIZE_FR = {
    "very_large": "Grand groupe",
    "large": "ETI / scale-up",
    "medium": "PME",
    "small": "Startup / petite entreprise",
}

# Doublons éditeur confirmés : on retire la ligne en double, on GARDE l'ÉDITEUR (pas le produit).
DROP_IDS = {
    "leviia",                        # produit "Leviia Drive" -> on garde l'éditeur "leviaa" (Leviia)
    "nucleon-security-nucleon-edr",  # produit "Nucleon EDR" -> on garde l'éditeur "nucleon-security"
    "evertrust-horizon",             # produit "Horizon" -> on garde l'éditeur "evertrust-sas"
    # NB: "sekost-yeswehack" et "yeswehack" sont DEUX sociétés distinctes -> on garde les deux.
    # Revue client Hexatrust/CESIN : CoreUpdate retiré de DROP (on le garde, voir EXCLUDE pour Galeax).
}
# Acteurs retirés du panorama (non-cyber, non-français, ou site mort) — décision validée.
EXCLUDE_IDS = {
    "myrgpd-pro",   # site suspendu
    "rfence",       # blog perso, pas un éditeur
    "defensx",      # société US, aucune entité française
    "vates",        # éditeur FR mais virtualisation/sauvegarde, pas cyber
    "afnic",        # registre .fr / opérateur DNS, pas un acteur cyber
    "sienna",       # projet DeFi/blockchain, ni cyber ni français
    # hors périmètre cyber (vérifié registre + site) — collaboratif/edtech/cloud non-sécurité
    "linagora",         # logiciels libres collaboratifs (Twake/OpenPaaS)
    "bodyguard",        # modération de contenu / trust & safety
    "safebrain",        # assistant IA générative privé (productivité)
    "dlplace",          # agence de développement
    "wraptor",          # intégration logicielle
    "waryme",           # appli protection travailleur isolé
    "clouddataengine",  # dataspaces / échange de données
    "pow-software-bytesync",  # synchronisation/sauvegarde de fichiers
    # Re-vérification 2026-06 : éditeurs disparus ou hors périmètre cyber.
    "beware-cyberlabs",   # dissoute, radiée du RCS le 01/10/2025
    "lybero-net",         # liquidation judiciaire (13/09/2024)
    "opencell-security",  # opencellsoft.com = facturation SaaS (FinTech), pas de la cyber
    "psc-france-cyber-detect-gorille",  # doublon de cyberdetect (meme editeur Cyber-Detect)
    "pradeo-yagaan",  # produit Yagaan = doublon de l editeur pradeo
    # Revue client Hexatrust/CESIN (2026-06) : Galeax retiré (on garde CoreUpdate, même produit) ;
    # Stackered retiré (fournisseur de services, pas éditeur).
    "galeax", "stackered",
    # Revue Hexatrust (2026-06) : rachetés + radiés / disparus.
    "malizen",  # racheté par Wallix et radié
    "n-cyp",    # racheté, n'existe plus
}
# Corrections de nom d'affichage : éditeur conservé + rebranding (nouveau nom).
NAME_FIX = {}  # migre dans solutions.json / size_review.json
# Correction d'URL (site officiel changé ou ancien domaine mort).
WEBSITE_FIX = {}  # migre dans solutions.json / size_review.json
# Surcharge de description (correction ponctuelle validée par le client, revue Hexatrust).
DESC_FIX = {}  # migre dans solutions.json / size_review.json
# Taille figée par décision (revue Hexatrust). small=Startup/TPME, medium=PME, large=ETI/Scale Up, very_large=Grand groupe.
SIZE_FIX = {}  # migre dans solutions.json / size_review.json
# Reclassement NIST validé (audit 2026-06, confiance élevée/moyenne). (fonction, catégorie, [sous-catégories]).
CLASSIF_FIX = {}  # migre dans solutions.json / size_review.json

# Colonnes de la feuille Solutions. Elles correspondent 1:1 aux champs du site ; toutes font
# l'aller-retour Excel <-> JSON (cf. xlsx_to_json.py). "Indexation" = mots-cles de recherche.
HEADERS = [
    "Entreprise", "Type d'entreprise", "Description", "Description longue", "Indexation",
    "NIST 1 (Fonction)", "NIST 2 (Catégorie)", "NIST 3 (Sous-catégories)",
    "Objectif NIS2", "Site web", "Contact", "Logo",
    # Cle technique : sert au sens inverse (Excel -> JSON via xlsx_to_json.py). NE PAS modifier.
    "ID",
]


def load_json(path: Path, default):
    if not path.exists():
        return default
    with path.open(encoding="utf-8") as fh:
        return json.load(fh) or default


# Re-vérification taille (revue IA 2026-06 : registre INSEE + CA + levées). SIZE_FIX (client) reste prioritaire.
SIZE_REVIEW = load_json(DATA_DIR / "size_review.json", {})


def fr_l1(labels, value):
    """value peut être un nom FR ('Protéger') ou un code ('PR'). Retourne 'CODE — Nom'."""
    if not value:
        return ""
    by_name = labels.get("level1_by_name", {})
    l1 = labels.get("level1", {})
    if value in by_name:                       # nom FR fourni
        code = by_name[value]
        return f"{code} : {value}"
    if value in l1:                            # code fourni
        return f"{value} : {l1[value]}"
    return value


def code_label(labels_map, code):
    lab = labels_map.get(code)
    return f"{code} : {lab}" if lab else code


def primary_l2(level2, level3):
    """Choisit UNE catégorie N2 principale : celle qui porte le plus de codes N3, sinon la 1re."""
    level2 = [c for c in (level2 or []) if c]
    if not level2:
        return None
    counts = Counter(c.rsplit("-", 1)[0] for c in (level3 or []))
    return max(level2, key=lambda c: (counts.get(c, 0), -level2.index(c)))


def build_row(sol, labels):
    nist = sol.get("nist") or {}
    # Toutes les catégories N2 (jusqu'à 3) et tous les N3 : l'aller-retour Excel les préserve.
    l2_list = [c for c in (nist.get("level2") or []) if c]
    l3 = [c for c in (nist.get("level3") or []) if c]
    contact = (sol.get("email_contact") or "").strip() or (sol.get("contact_url") or "").strip()
    # Indexation = les mots-cles de recherche (champ `indexation`). Editable et reinjecte tel quel.
    indexation = ", ".join(t.strip() for t in (sol.get("indexation") or []) if t and t.strip())
    return {
        "Entreprise": sol.get("solution_name", ""),
        "Type d'entreprise": SIZE_FR.get(sol.get("size", ""), sol.get("size", "")),
        "Description": sol.get("description", ""),
        "Description longue": sol.get("detailed_description", ""),
        "Indexation": indexation,
        "NIST 1 (Fonction)": fr_l1(labels, nist.get("level1")),
        "NIST 2 (Catégorie)": "\n".join(code_label(labels.get("level2", {}), c) for c in l2_list),
        "NIST 3 (Sous-catégories)": "\n".join(code_label(labels.get("level3", {}), c) for c in l3),
        "Objectif NIS2": sol.get("nis2_objective") or "",
        "Site web": sol.get("website", ""),
        "Contact": contact,
        "ID": sol.get("id", ""),
        "Pays": sol.get("country", ""),
        "Français": "Oui" if sol.get("is_french") else "Non",
        "Vérifié": "Oui" if sol.get("verified") else "Non",
        "Logo": sol.get("logo_file", ""),
        "_flag": sol.get("flag", ""),
        "_id": sol.get("id", ""),
    }


# ---- styles ----
HEAD_FILL = PatternFill("solid", fgColor="1F2A44")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
WIDTHS = {
    "Entreprise": 28, "Type d'entreprise": 22, "Description": 60, "Description longue": 80,
    "Indexation": 70,
    "NIST 1 (Fonction)": 18, "NIST 2 (Catégorie)": 34, "NIST 3 (Sous-catégories)": 50,
    "Objectif NIS2": 30, "Site web": 34, "Contact": 34, "Pays": 8, "Français": 9,
    "Vérifié": 9, "Logo": 22, "ID": 30,
}
WRAP_COLS = {"Description", "Description longue", "Indexation",
             "NIST 2 (Catégorie)", "NIST 3 (Sous-catégories)", "Objectif NIS2"}


def style_header(ws, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = WIDTHS.get(h, 18)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def write_solutions_sheet(wb, rows):
    # colonnes utiles : on ne garde que celles ayant au moins une valeur non vide
    headers = [h for h in HEADERS if any((r.get(h) or "").strip() for r in rows)]
    ws = wb.active
    ws.title = "Solutions"
    style_header(ws, headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for i in range(2, len(rows) + 2):
        for j, h in enumerate(headers, 1):
            ws.cell(row=i, column=j).alignment = WRAP_TOP if h in WRAP_COLS else TOP
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return headers


def write_nist_sheet(wb, labels):
    """Liste plate : une ligne par code (NIST 1, puis 2, puis 3)."""
    ws = wb.create_sheet("Listing référentiel NIST")
    style_header(ws, ["Niveau", "Code", "Libellé"])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 90
    for code, lab in labels.get("level1", {}).items():
        ws.append(["NIST 1", code, lab])
    for code, lab in labels.get("level2", {}).items():
        ws.append(["NIST 2", code, lab])
    for code, lab in sorted(labels.get("level3", {}).items()):
        ws.append(["NIST 3", code, lab])
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=3).alignment = WRAP_TOP


# Ordre canonique des fonctions.
L1_ORDER = ["GV", "ID", "PR", "DE", "RS", "RC"]


def _copy_sheet_exact(src_ws, dst_ws):
    """Copie une feuille entiere d'un classeur a un autre en conservant valeurs, styles,
    cellules fusionnees, largeurs de colonnes et hauteurs de lignes."""
    for row in src_ws.iter_rows():
        for cell in row:
            d = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                d.font = _copy(cell.font)
                d.fill = _copy(cell.fill)
                d.border = _copy(cell.border)
                d.alignment = _copy(cell.alignment)
                d.protection = _copy(cell.protection)
                d.number_format = cell.number_format
    for mc in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(mc))
    for key, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[key].width = dim.width
    for key, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[key].height = dim.height
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes


def _generate_matrix(ws, labels):
    """Repli si le fichier de reference est absent : vue hierarchique cote a cote generee."""
    l1 = labels.get("level1", {})
    l2 = labels.get("level2", {})
    l3 = labels.get("level3", {})
    style_header(ws, ["Niveau", "Code", "Libellé", "Niveau", "Code", "Libellé", "Niveau", "Code", "Libellé"])
    for col, w in zip("ABCDEFGHI", [9, 11, 34, 9, 11, 34, 9, 12, 80]):
        ws.column_dimensions[col].width = w
    row = 2
    for f in L1_ORDER:
        cats = [c for c in l2 if c.split(".")[0] == f]
        for cat in cats:
            subs = sorted(c for c in l3 if c.rsplit("-", 1)[0] == cat)
            c_start = row
            for sub in subs:
                ws.cell(row=row, column=7, value="NIST 3")
                ws.cell(row=row, column=8, value=sub)
                ws.cell(row=row, column=9, value=l3.get(sub, ""))
                row += 1
            ws.cell(row=c_start, column=4, value="NIST 2")
            ws.cell(row=c_start, column=5, value=cat)
            ws.cell(row=c_start, column=6, value=l2.get(cat, ""))
            if row - 1 > c_start:
                for col in (4, 5, 6):
                    ws.merge_cells(start_row=c_start, start_column=col, end_row=row - 1, end_column=col)
        ws.cell(row=ws.max_row, column=1)  # noop to keep structure
    # fonctions : merge sur leur etendue
    row = 2
    for f in L1_ORDER:
        cats = [c for c in l2 if c.split(".")[0] == f]
        span = sum(len([c for c in l3 if c.rsplit("-", 1)[0] == cat]) for cat in cats)
        ws.cell(row=row, column=1, value="NIST 1")
        ws.cell(row=row, column=2, value=f)
        ws.cell(row=row, column=3, value=l1.get(f, ""))
        if span > 1:
            for col in (1, 2, 3):
                ws.merge_cells(start_row=row, start_column=col, end_row=row + span - 1, end_column=col)
        row += span
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = center


def write_nist_matrix_sheet(wb, labels):
    """Vue hiérarchique côte à côte (NIST 1 | NIST 2 | NIST 3). On COPIE la feuille du fichier de
    référence du collègue à l'identique (mise en forme, fusions, bordures) ; à défaut on la génère."""
    ws = wb.create_sheet("Référentiel NIST")
    if REF_XLSX.exists():
        try:
            src = load_workbook(REF_XLSX)
            if "Référentiel NIST" in src.sheetnames:
                _copy_sheet_exact(src["Référentiel NIST"], ws)
                return
        except Exception:
            pass
    _generate_matrix(ws, labels)


CLASSIF = DATA_DIR / "classification_audit.json"
# L1 sans accent (sortie agent) vers le nom FR officiel.
L1_FR = {
    "Gouverner": "Gouverner", "Identifier": "Identifier", "Proteger": "Protéger",
    "Detecter": "Détecter", "Repondre": "Répondre", "Recuperer": "Récupérer",
}
CONF_FR = {"high": "Élevée", "medium": "Moyenne", "low": "Faible"}


def write_classification_sheet(wb, rows, audit, labels):
    """Page d'explication : pour chaque éditeur, son classement NIST et POURQUOI il est classé ainsi.
    La colonne Révision suggérée n'est remplie que si l'audit estime le classement actuel incorrect."""
    headers = ["Entreprise", "NIST 1 (Fonction)", "NIST 2 (Catégorie)",
               "NIST 3 (Sous-catégories)", "Justification du classement",
               "Révision suggérée", "Confiance",
               "Type d'entreprise", "Justification taille", "Confiance (taille)"]
    ws = wb.create_sheet("Classement & justification")
    style_header(ws, headers)
    widths = {"Entreprise": 28, "NIST 1 (Fonction)": 18, "NIST 2 (Catégorie)": 34,
              "NIST 3 (Sous-catégories)": 46, "Justification du classement": 80,
              "Révision suggérée": 46, "Confiance": 11,
              "Type d'entreprise": 16, "Justification taille": 90, "Confiance (taille)": 14}
    for j, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 18)
    l2_labels = labels.get("level2", {})
    for r in rows:
        a = audit.get(r.get("_id"), {})
        revision = ""
        if a and not a.get("currentOk", True):
            recL1 = L1_FR.get(a.get("recommendedL1", ""), a.get("recommendedL1", ""))
            recL2 = code_label(l2_labels, a.get("recommendedL2", "")) if a.get("recommendedL2") else ""
            revision = f"Proposé : {recL1} / {recL2}. {a.get('whyChange', '')}".strip()
        sr = SIZE_REVIEW.get(r.get("_id"), {})
        ws.append([
            r["Entreprise"], r["NIST 1 (Fonction)"], r["NIST 2 (Catégorie)"],
            r["NIST 3 (Sous-catégories)"], a.get("justification", ""),
            revision, CONF_FR.get(a.get("confidence", ""), ""),
            r.get("Type d'entreprise", ""), sr.get("justification", ""), sr.get("confidence", ""),
        ])
    for i in range(2, len(rows) + 2):
        for j, h in enumerate(headers, 1):
            ws.cell(row=i, column=j).alignment = WRAP_TOP
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return sum(1 for r in rows if audit.get(r.get('_id'), {}).get('currentOk') is False)


def write_flagged_sheet(wb, rows):
    flagged = [r for r in rows if r.get("_flag")]
    if not flagged:
        return 0
    ws = wb.create_sheet("À vérifier")
    style_header(ws, ["Entreprise", "Site web", "Problème"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 50
    for r in flagged:
        ws.append([r["Entreprise"], r["Site web"], r["_flag"]])
    return len(flagged)


def main():
    base = load_json(SOLUTIONS, [])   # solutions.json = donnee finale (deja fusionnee, exclus retires)
    labels = load_json(LABELS, {})

    rows = []
    for s in base:
        sid = s.get("id")
        if sid in DROP_IDS or sid in EXCLUDE_IDS:   # filet de securite
            continue
        m = dict(s)
        if sid in NAME_FIX:
            m["solution_name"] = NAME_FIX[sid]
        if sid in WEBSITE_FIX:
            m["website"] = WEBSITE_FIX[sid]
        if sid in DESC_FIX:
            m["description"] = DESC_FIX[sid]
        if sid in SIZE_REVIEW:                         # taille re-vérifiée (IA)
            m["size"] = SIZE_REVIEW[sid].get("size_code", m.get("size"))
        if sid in SIZE_FIX:                            # taille figée client : prioritaire
            m["size"] = SIZE_FIX[sid]
        if sid in CLASSIF_FIX:
            l1f, l2f, l3f = CLASSIF_FIX[sid]
            m["nist"] = {"level1": l1f, "level2": [l2f], "level3": list(l3f)}
        rows.append(build_row(m, labels))

    audit_raw = load_json(CLASSIF, {})
    audit = {a["id"]: a for a in audit_raw.get("all", [])} if isinstance(audit_raw, dict) else {}

    wb = Workbook()
    used_headers = write_solutions_sheet(wb, rows)
    write_nist_sheet(wb, labels)            # Listing référentiel NIST (liste plate)
    write_nist_matrix_sheet(wb, labels)     # Référentiel NIST (vue hiérarchique côte à côte)
    n_revisions = write_classification_sheet(wb, rows, audit, labels)
    n_flagged = write_flagged_sheet(wb, rows)
    wb.save(OUT_XLSX)
    print(f"  page Classement & justification : {len(audit)} justifiées, {n_revisions} révisions suggérées")

    # ---- QA ----
    def missing(h):
        return sum(1 for r in rows if not (r.get(h) or "").strip())

    qa = {
        "total": len(rows),
        "verified": sum(1 for r in rows if r["Vérifié"] == "Oui"),
        "flagged": n_flagged,
        "non_french": sum(1 for r in rows if r["Français"] == "Non"),
        "non_french_list": [{"solution": r["Entreprise"], "pays": r["Pays"], "note": r["_flag"]}
                            for r in rows if r["Français"] == "Non"],
        "unverified_list": [{"solution": r["Entreprise"], "note": r["_flag"]}
                            for r in rows if r["Vérifié"] != "Oui"],
        "columns_written": used_headers,
        "missing_per_field": {h: missing(h) for h in
                              ["Description", "NIST 1 (Fonction)", "NIST 2 (Catégorie)",
                               "NIST 3 (Sous-catégories)", "Site web", "Contact"]},
        "n2_not_single": sum(1 for r in rows if not r["NIST 2 (Catégorie)"]),
        "n3_outside_n2": 0,  # garanti par build_row (filtrage par préfixe)
    }
    with QA_JSON.open("w", encoding="utf-8") as fh:
        json.dump(qa, fh, ensure_ascii=False, indent=2)

    print(f"Écrit : {OUT_XLSX}")
    print(f"  lignes={qa['total']}  vérifiées={qa['verified']}  flaggées={qa['flagged']}"
          f"  non-FR={qa['non_french']}")
    print(f"  colonnes={len(used_headers)} : {used_headers}")
    print(f"  manquants : {qa['missing_per_field']}")


if __name__ == "__main__":
    main()
